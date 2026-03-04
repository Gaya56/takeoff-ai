"""
TakeoffPipeline: AI-powered construction takeoff from architectural PDFs.
Uses Gemini Vision to detect wall types and segments from floor plans.
"""

import os
import re
import json
import fitz  # PyMuPDF
import cv2
import numpy as np
from concurrent.futures import ThreadPoolExecutor, as_completed
from pydantic import BaseModel, Field
from typing import Optional, List
from google import genai
from google.genai import types
from scipy.spatial import cKDTree


# ─── Pydantic models for Gemini structured output ────────────────────────────

class WallType(BaseModel):
    code: str = Field(description="Element code, e.g. M1, C1, DOOR, WIN")
    description: str = Field(description="Human-readable description")
    category: str = Field(description="wall|door|window|structural|dimension|room|other")


class Detection(BaseModel):
    type_code: str = Field(description="Code from wall_types or standard codes")
    category: str = Field(default="wall", description="wall|door|window|structural|other")
    page: int = Field(default=0, ge=0)
    x1: float = Field(description="Start X as fraction of image width", ge=0.0, le=1.0)
    y1: float = Field(description="Start Y as fraction of image height", ge=0.0, le=1.0)
    x2: float = Field(description="End X as fraction of image width", ge=0.0, le=1.0)
    y2: float = Field(description="End Y as fraction of image height", ge=0.0, le=1.0)
    real_length_ft: float = Field(default=0.0, ge=0.0, description="Real-world length in feet")
    label: str = Field(default="", description="Short display label")
    confidence: float = Field(default=0.5, ge=0.0, le=1.0, description="Detection confidence")


class FloorPlanAnalysis(BaseModel):
    wall_types: List[WallType] = Field(default_factory=list)
    scale: Optional[str] = Field(default=None)
    detections: List[Detection] = Field(default_factory=list)


class FloorPlanBBox(BaseModel):
    floor_plan_x0: float = Field(ge=0.0, le=1.0, description="Left edge as fraction of image width")
    floor_plan_y0: float = Field(ge=0.0, le=1.0, description="Top edge as fraction of image height")
    floor_plan_x1: float = Field(ge=0.0, le=1.0, description="Right edge as fraction of image width")
    floor_plan_y1: float = Field(ge=0.0, le=1.0, description="Bottom edge as fraction of image height")
    scale_bar_pixels: float = Field(default=0.0, ge=0.0, description="Length of scale bar in image pixels")
    scale_bar_real_ft: float = Field(default=0.0, ge=0.0, description="Real-world length the scale bar represents in feet")


class LegendWallType(BaseModel):
    code: str = Field(description="Wall type code, e.g. M1, C1, DOOR, WIN")
    description: str = Field(description="Human-readable description")
    category: str = Field(description="wall|door|window|structural|dimension|room|other")
    material: str = Field(default="", description="Material composition, e.g. 'brick veneer + steel stud'")
    thickness_in: float = Field(default=0.0, description="Typical wall thickness in inches")


class LegendExtraction(BaseModel):
    wall_types: List[LegendWallType] = Field(default_factory=list)
    scale: Optional[str] = Field(default=None, description="Drawing scale, e.g. 1/8 inch = 1 ft-0 in")
    floor_label: Optional[str] = Field(default=None, description="Floor or story label")
    orientation: Optional[str] = Field(default=None, description="North arrow direction")


# ─── Color maps ───────────────────────────────────────────────────────────────

COLOR_MAP = {
    "M1": (0.9, 0.1, 0.1),    # Red — Brick exterior
    "M2": (0.1, 0.3, 0.9),    # Blue — Siding exterior
    "M3": (0.9, 0.5, 0.0),    # Orange — Other masonry
    "C1": (0.0, 0.6, 0.2),    # Green — Load-bearing partition
    "C2": (0.5, 0.0, 0.7),    # Purple — Stairway/fire partition
    "C3": (0.8, 0.8, 0.0),    # Yellow — Double/party wall
    "DOOR":  (0.0, 0.7, 0.8), # Cyan
    "WIN":   (0.0, 0.8, 0.6), # Teal
    "BEAM":  (0.6, 0.3, 0.0), # Brown
    "COL":   (0.4, 0.4, 0.4), # Gray
    "STAIR": (0.7, 0.5, 0.0), # Dark yellow
    "DIM":   (0.5, 0.5, 0.5), # Gray
    "ROOM":  (0.3, 0.6, 0.9), # Light blue
}

CATEGORY_COLOR = {
    "wall":       (0.7, 0.2, 0.2),
    "door":       (0.0, 0.7, 0.8),
    "window":     (0.0, 0.8, 0.6),
    "structural": (0.5, 0.3, 0.0),
    "dimension":  (0.5, 0.5, 0.5),
    "room":       (0.3, 0.6, 0.9),
    "other":      (0.6, 0.6, 0.6),
}

FLAGGED_COLOR = (1, 0.4, 0.5)
CONFIDENCE_THRESHOLD = 0.40

# Fallback crop fractions — used only if dynamic localization fails
FLOOR_PLAN_TOP_FRAC    = 0.05  # skip top 5% (conservative — keep almost everything)
FLOOR_PLAN_BOTTOM_FRAC = 0.95  # skip bottom 5% (title block)


# ─── Prompts ──────────────────────────────────────────────────────────────────

SYSTEM_INSTRUCTION = """You are a senior Canadian construction estimator with 20+ years reading architectural floor plans for material takeoff.

DOMAIN KNOWLEDGE — Canadian Wall Type Codes:
- M1: Brick veneer exterior wall (typically 4" brick + air gap + sheathing + steel/wood stud)
- M2: Siding exterior wall (vinyl/Canexel/fiber cement + sheathing + stud)
- M3: Other masonry exterior (stone veneer, stucco, precast)
- C1: Load-bearing interior partition (steel stud, typically 3-5/8" or 6")
- C2: Fire-rated partition / stairwell enclosure (2-hour rated, double layer gypsum)
- C3: Double/party wall (two separate stud walls with air gap, sound-rated)
- DOOR: Door opening (wall break + swing arc in plan view)
- WIN: Window opening (wall break + parallel lines or X pattern)
- BEAM: Structural beam (dashed or bold line, often with size callout like W12x26)
- COL: Column (filled/hollow square or circle, typically 12"–24")
- STAIR: Stairway (parallel lines with arrow indicating direction)

MEASUREMENT METHODOLOGY:
- Measure wall segments CENTERLINE-to-CENTERLINE
- Segments BREAK at: openings (doors/windows), corners, material type changes, intersecting walls
- A 40-foot wall with 2 doors creates 3 wall segments + 2 door detections = 5 total detections
- Standard reference dimensions: residential door = 2'-10" to 3'-0" wide, double door = 5'-0" to 6'-0", standard window = 3'-0" to 4'-0"

CAD SYMBOL CONVENTIONS (plan view):
- Doors: gap in wall line + quarter-circle arc showing swing direction
- Windows: gap in wall line + thin parallel lines or X inside gap
- Dimension lines: thin lines with arrows/ticks at endpoints + number — SKIP THESE (they are NOT walls)
- Section cut markers: circle with number/letter — SKIP THESE
- Hatch patterns: diagonal lines inside walls showing material — SKIP THESE (they indicate fill, not separate elements)
- Room labels: text centered in room — these are ROOM type, text label only

COORDINATE CONVENTIONS:
- (0.0, 0.0) = TOP-LEFT corner of the image
- (1.0, 1.0) = BOTTOM-RIGHT corner of the image
- x1,y1 = START endpoint of segment centerline; x2,y2 = END endpoint
- All values MUST be between 0.0 and 1.0 inclusive
- Each detection = ONE wall segment or ONE opening — never an entire room outline
- Horizontal walls: y1 ≈ y2. Vertical walls: x1 ≈ x2."""


GEMINI_PROMPT_V2 = """Analyze this architectural floor plan image. The image shows the main drawing area — legend and title block have been cropped out.

STAGE 1 — VERIFY SCALE:
Cross-check the provided scale against a known reference element visible in the drawing.
Standard references: residential door opening ≈ 3'-0" wide, hallway ≈ 3'-6" to 4'-0" wide, parking stall ≈ 9'-0" wide.
If scale seems wrong, note it and proceed with best estimate.

STAGE 2 — SYSTEMATIC SCAN:
Scan the floor plan from TOP-LEFT to BOTTOM-RIGHT, row by row.
For EVERY wall segment, door, window, and structural element:
- Identify its type code from the legend
- Mark x1,y1 (start) and x2,y2 (end) as fractions of image dimensions
- Estimate real_length_ft using the drawing scale
- Assign a confidence score (0.0–1.0)

SEGMENT ISOLATION RULES (critical for accuracy):
- Where a door or window interrupts a wall, the wall is broken into SEPARATE segments on each side of the opening
- Where two different wall types meet, each gets its own detection
- At corners, each leg is a separate segment
- T-junctions: the continuous wall is one segment, the branch is another
- Example: 40ft wall with 2 doors → detect as: WALL(~15ft) + DOOR(3ft) + WALL(~9ft) + DOOR(3ft) + WALL(~10ft)

STAGE 3 — VALIDATE:
After scanning, verify:
- Total detections should be 40–150 for a typical floor plan page
- Every visible room should have walls on all sides
- Wall types are consistent with the provided legend
- No detection spans more than 50ft unless it's a BEAM

DO NOT DETECT:
- Dimension lines (thin lines with arrows and numbers)
- Text labels or room names (these are ROOM type detections with text only)
- Section cut markers (circles with letters/numbers)
- Hatch patterns (diagonal fill lines inside walls)
- Title block elements
- North arrow or scale bar graphics

BE EXHAUSTIVE. Missing a wall is worse than including an uncertain one with low confidence."""


LOCALIZE_PROMPT = """You are looking at a full architectural drawing sheet.

Find TWO things:

1. FLOOR PLAN BOUNDING BOX — the region containing the main floor plan geometry (rooms, walls, doors, windows).
Exclude: legend panel, ASSEMBLAGES TYPIQUES / assembly details (cross-sections), title block, north arrow, scale bar border.

2. SCALE BAR — if a graphic scale bar is visible, measure:
- Its length in pixels (approximate)
- What real-world distance it represents in feet

Return all values as decimal fractions of total image dimensions.
Return valid JSON only. No markdown."""


LEGEND_EXTRACTION_PROMPT = """Read the legend/assemblies panel on this architectural drawing sheet.

Extract ALL of the following:

1. WALL TYPE CODES — Every code visible (M1, M2, M3, C1, C2, C3, DOOR, WIN, BEAM, COL, STAIR, etc.):
   - code: the short code
   - description: the full text description
   - category: wall|door|window|structural|dimension|room|other
   - material: material composition if described (e.g. "brick veneer + steel stud")
   - thickness_in: wall thickness in inches if shown

2. SCALE — the drawing scale if visible (e.g. "1/8 inch = 1 ft-0 in")

3. FLOOR LABEL — which floor/story this sheet represents (e.g. "Rez-de-chaussée", "1st Floor", "Level 2")

4. ORIENTATION — north arrow direction if visible (e.g. "North is up", "North is to the right")

Be thorough — include every code even if the description is partially obscured."""


# ─── Image preprocessing ─────────────────────────────────────────────────────

def preprocess_floor_plan_image(img_bytes: bytes) -> bytes:
    """Remove bright colored room fills while preserving wall line colors."""
    nparr = np.frombuffer(img_bytes, np.uint8)
    img = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
    if img is None:
        return img_bytes

    # Remove colored room fills but keep wall colors
    # Room fills: high saturation + high value (bright pastel colors)
    # Wall lines: low value (dark lines regardless of hue) — preserve these
    hsv = cv2.cvtColor(img, cv2.COLOR_BGR2HSV)
    saturation = hsv[:, :, 1]
    value = hsv[:, :, 2]
    fill_mask = (saturation > 60) & (value > 150)
    img[fill_mask] = [255, 255, 255]

    # Sharpen edges without converting to grayscale
    kernel = np.array([[-1, -1, -1], [-1, 9, -1], [-1, -1, -1]])
    img = cv2.filter2D(img, -1, kernel)

    _, buffer = cv2.imencode('.png', img)
    return buffer.tobytes()


# ─── Coordinate helpers ───────────────────────────────────────────────────────

def _is_bleed_line(x1: float, y1: float, x2: float, y2: float,
                   pw: float, ph: float) -> bool:
    """Return True if the line looks like a page-edge bleed artifact."""
    margin = 5.0
    near_edge = (
        x1 <= margin or x2 <= margin or
        x1 >= pw - margin or x2 >= pw - margin or
        y1 <= margin or y2 <= margin or
        y1 >= ph - margin or y2 >= ph - margin
    )
    span_x = abs(x2 - x1)
    span_y = abs(y2 - y1)
    is_large = span_x > pw * 0.70 or span_y > ph * 0.70
    return is_large and near_edge


def _is_percentage_coords(det: dict) -> bool:
    """Return True if detection coordinates look like fractions (0–1)."""
    vals = [det.get("x1", 0), det.get("y1", 0), det.get("x2", 1), det.get("y2", 1)]
    return all(0.0 <= v <= 1.0 for v in vals)


def _parse_json_safe(text: str) -> dict:
    """Parse JSON from Gemini response, handling common formatting issues."""
    text = text.strip()
    if text.startswith("```"):
        lines = text.split("\n")[1:]
        if lines and lines[-1].strip() == "```":
            lines = lines[:-1]
        text = "\n".join(lines)

    try:
        parsed = json.loads(text)
        if isinstance(parsed, dict):
            return parsed
        if isinstance(parsed, list):
            return {"wall_types": [], "scale": None, "detections": parsed}
    except json.JSONDecodeError:
        pass

    start = text.find("{")
    end = text.rfind("}") + 1
    if start >= 0 and end > start:
        try:
            return json.loads(text[start:end])
        except json.JSONDecodeError:
            pass

    cleaned = re.sub(r',\s*([}\]])', r'\1', text[start:end] if start >= 0 else text)
    try:
        return json.loads(cleaned)
    except json.JSONDecodeError:
        pass

    return {"wall_types": [], "scale": None, "detections": [], "_parse_error": True}


# ─── Pipeline ─────────────────────────────────────────────────────────────────

class TakeoffPipeline:
    def __init__(self, gemini_api_key: str = None):
        self.api_key = gemini_api_key or os.environ.get("GEMINI_API_KEY")
        if not self.api_key:
            raise ValueError("GEMINI_API_KEY is required")
        self.client = genai.Client(api_key=self.api_key)
        self.pdf_path = None
        self.doc = None
        self.analysis = None
        self._legend_data = None
        self._errors = []
        self._page_vectors = {}
        self._snap_trees = {}

    def ingest_pdf(self, pdf_path: str) -> dict:
        """Load PDF and extract page info."""
        self.pdf_path = pdf_path
        self.doc = fitz.open(pdf_path)
        info = {
            "path": pdf_path,
            "page_count": len(self.doc),
            "pages": [],
            "is_vector": False,
        }
        vector_path_count = 0
        for i, page in enumerate(self.doc):
            rect = page.rect
            paths = page.get_drawings()
            vector_path_count += len(paths)
            info["pages"].append({
                "page": i,
                "width": rect.width,
                "height": rect.height,
                "vector_paths": len(paths),
            })
        info["is_vector"] = vector_path_count > 50
        return info

    # ─── Phase 1: Vector line extraction + spatial index ─────────────────────

    def _extract_vector_lines(self, page_idx: int) -> list:
        """Extract wall-like vector line segments from PDF page using PyMuPDF."""
        page = self.doc[page_idx]
        paths = page.get_drawings()
        lines = []
        for path in paths:
            for item in path["items"]:
                if item[0] == "l":  # line segment
                    p1, p2 = item[1], item[2]
                    length = ((p2.x - p1.x)**2 + (p2.y - p1.y)**2)**0.5
                    if length > 5.0:  # skip tiny artifacts
                        lines.append({
                            "x1": p1.x, "y1": p1.y,
                            "x2": p2.x, "y2": p2.y,
                            "width": path.get("width", 1.0),
                            "color": path.get("color"),
                            "length": length,
                        })
        return lines

    def _build_snap_index(self, vector_lines: list) -> tuple:
        """Build a cKDTree spatial index from vector line endpoints for O(log n) snapping."""
        if not vector_lines:
            return None, []
        endpoints = []
        for line in vector_lines:
            endpoints.append([line["x1"], line["y1"]])
            endpoints.append([line["x2"], line["y2"]])
        ep_array = np.array(endpoints)
        tree = cKDTree(ep_array)
        return tree, endpoints

    def _snap_to_vectors(self, x1, y1, x2, y2, tree, endpoints, threshold=15.0):
        """Snap detection endpoints to nearest PDF vector line endpoints."""
        if tree is None:
            return x1, y1, x2, y2
        dist1, idx1 = tree.query([x1, y1])
        dist2, idx2 = tree.query([x2, y2])
        if dist1 < threshold:
            x1, y1 = endpoints[idx1]
        if dist2 < threshold:
            x2, y2 = endpoints[idx2]
        return x1, y1, x2, y2

    # ─── Phase 2: Legend extraction (Call 1) ──────────────────────────────────

    def _extract_legend(self, page_idx: int = 0) -> dict:
        """Call 1: Use Gemini Flash to extract legend/wall types from drawing sheet."""
        page = self.doc[page_idx]
        mat = fitz.Matrix(1.5, 1.5)
        pix = page.get_pixmap(matrix=mat)
        img_bytes = pix.tobytes("png")

        try:
            response = self.client.models.generate_content(
                model="gemini-2.5-flash",
                contents=[
                    types.Content(parts=[
                        types.Part(
                            inline_data=types.Blob(mime_type="image/png", data=img_bytes)
                        ),
                        types.Part(text=LEGEND_EXTRACTION_PROMPT),
                    ])
                ],
                config=types.GenerateContentConfig(
                    temperature=0.0,
                    max_output_tokens=2048,
                    response_mime_type="application/json",
                    response_schema=LegendExtraction,
                ),
            )

            legend = response.parsed
            if legend is None:
                legend = LegendExtraction.model_validate_json(response.text)

            return {
                "wall_types": [wt.model_dump() for wt in legend.wall_types],
                "scale": legend.scale,
                "floor_label": legend.floor_label,
                "orientation": legend.orientation,
            }
        except Exception as e:
            self._errors.append({
                "page": page_idx,
                "error": f"Legend extraction failed: {e}",
                "stage": "legend",
            })
            return {"wall_types": [], "scale": None, "floor_label": None, "orientation": None}

    # ─── Phase 3: Localization (Call 2) ───────────────────────────────────────

    def _localize_floor_plan(self, page_idx: int) -> dict:
        """
        Call 2: Use Gemini Flash to find the floor plan bounding box + scale calibration.
        Returns dict with bbox list and scale info.
        Falls back to conservative static crop on failure.
        """
        page = self.doc[page_idx]
        mat = fitz.Matrix(1.5, 1.5)
        pix = page.get_pixmap(matrix=mat)
        img_bytes = pix.tobytes("png")

        try:
            response = self.client.models.generate_content(
                model="gemini-2.5-flash",
                contents=[
                    types.Content(parts=[
                        types.Part(
                            inline_data=types.Blob(mime_type="image/png", data=img_bytes)
                        ),
                        types.Part(text=LOCALIZE_PROMPT),
                    ])
                ],
                config=types.GenerateContentConfig(
                    temperature=0.0,
                    max_output_tokens=2048,
                    response_mime_type="application/json",
                    response_schema=FloorPlanBBox,
                ),
            )

            bbox = response.parsed
            if bbox is None:
                bbox = FloorPlanBBox.model_validate_json(response.text)

            if (bbox.floor_plan_x1 - bbox.floor_plan_x0 > 0.10 and
                    bbox.floor_plan_y1 - bbox.floor_plan_y0 > 0.10):
                return {
                    "bbox": [bbox.floor_plan_x0, bbox.floor_plan_y0,
                             bbox.floor_plan_x1, bbox.floor_plan_y1],
                    "scale_bar_pixels": bbox.scale_bar_pixels,
                    "scale_bar_real_ft": bbox.scale_bar_real_ft,
                    "fallback": False,
                }
        except Exception as e:
            self._errors.append({
                "page": page_idx,
                "error": f"Localization exception: {e}",
                "stage": "localize",
            })

        # Fallback: conservative static crop
        self._errors.append({
            "page": page_idx,
            "error": "Localization fell back to static crop",
            "stage": "localize",
        })
        return {
            "bbox": [0.0, FLOOR_PLAN_TOP_FRAC, 1.0, FLOOR_PLAN_BOTTOM_FRAC],
            "scale_bar_pixels": 0.0,
            "scale_bar_real_ft": 0.0,
            "fallback": True,
        }

    # ─── Phase 3: Detection (Call 3) ─────────────────────────────────────────

    def _analyze_single_page(self, page_idx: int, user_notes: str = "",
                             legend_data: dict = None, user_context: dict = None) -> dict:
        """3-call pipeline per page: localize -> detect (with legend context). Thread-safe."""
        page = self.doc[page_idx]
        pw = page.rect.width
        ph = page.rect.height

        # Call 2: Flash localization — find floor plan bbox + scale calibration
        loc = self._localize_floor_plan(page_idx)
        x0_f, y0_f, x1_f, y1_f = loc["bbox"]

        crop_rect = fitz.Rect(x0_f * pw, y0_f * ph, x1_f * pw, y1_f * ph)

        # Extract vector lines for snapping (Phase 1)
        vector_lines = self._extract_vector_lines(page_idx)
        tree, endpoints = self._build_snap_index(vector_lines)
        self._page_vectors[page_idx] = vector_lines
        self._snap_trees[page_idx] = (tree, endpoints)

        # Render cropped area at 3x zoom + preprocess
        mat = fitz.Matrix(3.0, 3.0)
        pix = page.get_pixmap(matrix=mat, clip=crop_rect)
        img_bytes = pix.tobytes("png")
        img_bytes = preprocess_floor_plan_image(img_bytes)

        # Build prompt with legend context (Phase 2)
        legend_section = ""
        if legend_data and legend_data.get("wall_types"):
            legend_section = f"\n\nLEGEND (pre-extracted from this drawing set):\n{json.dumps(legend_data['wall_types'], indent=2)}"

        scale_section = ""
        if legend_data and legend_data.get("scale"):
            scale_section = f"\n\nSCALE: {legend_data['scale']}"
        if loc.get("scale_bar_real_ft") and loc["scale_bar_real_ft"] > 0:
            scale_section += f"\nScale bar calibration: {loc['scale_bar_pixels']:.0f}px = {loc['scale_bar_real_ft']:.1f}ft"

        # Structured engineer context (Phase 6)
        context_section = ""
        if user_context:
            parts = []
            if user_context.get("building_type"):
                parts.append(f"Building type: {user_context['building_type']}")
            if user_context.get("wall_codes"):
                parts.append(f"Wall codes: {user_context['wall_codes']}")
            if user_context.get("scale"):
                parts.append(f"Engineer-provided scale: {user_context['scale']}")
                scale_section = f"\n\nSCALE (engineer-provided): {user_context['scale']}"
            if user_context.get("num_floors"):
                parts.append(f"Number of floors: {user_context['num_floors']}")
            if user_context.get("special_instructions"):
                parts.append(f"Special instructions: {user_context['special_instructions']}")
            if parts:
                context_section = "\n\nENGINEER CONTEXT:\n" + "\n".join(parts)

        notes_section = (
            f"\n\nCLIENT NOTES (pay special attention):\n{user_notes.strip()}"
            if user_notes and user_notes.strip() else ""
        )

        page_prompt = (
            f"This is page {page_idx + 1} of {len(self.doc)} of a Canadian architectural floor plan."
            f"{legend_section}{scale_section}{context_section}{notes_section}"
            f"\n\n{GEMINI_PROMPT_V2}"
        )

        try:
            response = self.client.models.generate_content(
                model="gemini-2.5-pro",
                contents=[
                    types.Content(parts=[
                        types.Part(
                            inline_data=types.Blob(mime_type="image/png", data=img_bytes)
                        ),
                        types.Part(text=page_prompt),
                    ])
                ],
                config=types.GenerateContentConfig(
                    system_instruction=SYSTEM_INSTRUCTION,
                    temperature=0.0,
                    max_output_tokens=65536,
                    response_mime_type="application/json",
                    response_schema=FloorPlanAnalysis,
                    media_resolution="MEDIA_RESOLUTION_HIGH",
                ),
            )

            analysis = response.parsed
            if analysis is None:
                try:
                    analysis = FloorPlanAnalysis.model_validate_json(response.text)
                except Exception:
                    raw = _parse_json_safe(response.text)
                    self._errors.append({
                        "page": page_idx,
                        "error": "JSON parse failure — fell back to manual parsing",
                        "raw_response_length": len(response.text),
                        "stage": "detect",
                    })
                    analysis = FloorPlanAnalysis(scale=raw.get("scale"))
                    for wt in raw.get("wall_types", []):
                        if isinstance(wt, dict):
                            try:
                                analysis.wall_types.append(WallType(**wt))
                            except Exception:
                                pass
                    for det in raw.get("detections", []):
                        if isinstance(det, dict):
                            try:
                                analysis.detections.append(Detection(**det))
                            except Exception:
                                pass
        except Exception as e:
            self._errors.append({
                "page": page_idx,
                "error": f"Gemini Pro detection call failed: {e}",
                "stage": "detect",
            })
            analysis = FloorPlanAnalysis()

        for det in analysis.detections:
            det.page = page_idx

        return {
            "wall_types": [wt.model_dump() for wt in analysis.wall_types],
            "scale": analysis.scale,
            "detections": [det.model_dump() for det in analysis.detections],
            "crop_rect": [crop_rect.x0, crop_rect.y0, crop_rect.x1, crop_rect.y1],
            "detection_count": len(analysis.detections),
            "localization_fallback": loc.get("fallback", False),
        }

    # ─── Phase 4: Post-detection validation ──────────────────────────────────

    def _validate_detections(self, detections: list, page_idx: int, scale_str: str = None, known_wall_codes: list = None) -> list:
        """Post-detection validation: dedup, geometric snap, scale check, type consistency."""
        if not detections:
            return detections

        # 4a. Duplicate removal — merge detections with >80% endpoint overlap
        cleaned = []
        for det in detections:
            is_dup = False
            for existing in cleaned:
                if existing.get("page") != det.get("page"):
                    continue
                d1 = ((det["x1"] - existing["x1"])**2 + (det["y1"] - existing["y1"])**2)**0.5
                d2 = ((det["x2"] - existing["x2"])**2 + (det["y2"] - existing["y2"])**2)**0.5
                # Also check reversed endpoints
                d3 = ((det["x1"] - existing["x2"])**2 + (det["y1"] - existing["y2"])**2)**0.5
                d4 = ((det["x2"] - existing["x1"])**2 + (det["y2"] - existing["y1"])**2)**0.5
                # Threshold in normalized coords (0-1 range): ~0.015 = ~10pt on 700pt page
                thresh = 0.015
                if (d1 < thresh and d2 < thresh) or (d3 < thresh and d4 < thresh):
                    # Keep higher confidence
                    if det.get("confidence", 0) > existing.get("confidence", 0):
                        existing.update(det)
                    is_dup = True
                    break
            if not is_dup:
                cleaned.append(det)

        # 4b. Scale verification
        if scale_str:
            scale_ratio = self._parse_scale_ratio(scale_str)
            if scale_ratio:
                for det in cleaned:
                    length = det.get("real_length_ft", 0)
                    if length > 200:
                        det["confidence"] = min(det.get("confidence", 0.5), 0.3)
                        det["label"] = det.get("label", "") + " [SCALE?]"
                    elif 0 < length < 0.3:
                        det["confidence"] = min(det.get("confidence", 0.5), 0.3)

        # 4c. Geometric constraints — snap near-axis lines
        for det in cleaned:
            cat = det.get("category", "wall")
            if cat in ("room", "dimension"):
                continue
            dx = det["x2"] - det["x1"]
            dy = det["y2"] - det["y1"]
            length = (dx**2 + dy**2)**0.5
            if length < 0.001:
                continue
            angle = abs(np.degrees(np.arctan2(dy, dx))) % 180
            # Snap to horizontal (0 deg or 180 deg)
            if angle < 5 or angle > 175:
                mid_y = (det["y1"] + det["y2"]) / 2
                det["y1"] = mid_y
                det["y2"] = mid_y
            # Snap to vertical (90 deg)
            elif 85 < angle < 95:
                mid_x = (det["x1"] + det["x2"]) / 2
                det["x1"] = mid_x
                det["x2"] = mid_x
            # Snap to 45 deg or 135 deg
            elif 40 < angle < 50 or 130 < angle < 140:
                # Keep diagonal but ensure exact 45 deg
                mid_x = (det["x1"] + det["x2"]) / 2
                mid_y = (det["y1"] + det["y2"]) / 2
                half_len = length / 2
                sign_x = 1 if dx >= 0 else -1
                sign_y = 1 if dy >= 0 else -1
                component = half_len / (2**0.5)
                det["x1"] = mid_x - sign_x * component
                det["y1"] = mid_y - sign_y * component
                det["x2"] = mid_x + sign_x * component
                det["y2"] = mid_y + sign_y * component

        # 4d. Wall type consistency check
        if known_wall_codes:
            valid_codes = set(known_wall_codes)
            # Always allow standard non-legend codes
            valid_codes.update({"DOOR", "WIN", "BEAM", "COL", "STAIR", "DIM", "ROOM", "UNKNOWN"})
            for det in cleaned:
                tc = det.get("type_code", "UNKNOWN")
                if tc not in valid_codes:
                    det["original_type_code"] = tc
                    det["type_code"] = "UNKNOWN"
                    det["confidence"] = min(det.get("confidence", 0.5), 0.4)

        return cleaned

    def _parse_scale_ratio(self, scale_str: str) -> float:
        """Parse a scale string like '1/8" = 1'-0"' into a numeric ratio (drawing units per real foot)."""
        if not scale_str:
            return 0.0
        try:
            # Try common formats: "1/8" = 1'-0"", "1:100", "1/4" = 1'-0""
            s = scale_str.replace('"', '').replace("'", '').replace('\\"', '').replace("\\'", '')
            # Format: "1/X = 1-0" (X inches = 1 foot)
            m = re.search(r'1\s*/\s*(\d+)\s*=\s*1', s)
            if m:
                return 1.0 / float(m.group(1))  # e.g., 1/8 = 0.125
            # Format: "1:NNN"
            m = re.search(r'1\s*:\s*(\d+)', s)
            if m:
                return 1.0 / float(m.group(1))
        except Exception:
            pass
        return 0.0

    # ─── Length computation from coordinates + scale ─────────────────────────

    def _compute_real_lengths(self, detections: list, page_crops: dict, scale_str: str):
        """Compute real_length_ft from normalized coords + crop rect + scale.

        Scale "1/8 inch = 1'-0 inch" means 1/8 inch on paper = 1 foot real.
        PDF points: 72 points = 1 inch.
        So points_per_foot = scale_fraction * 72 (e.g., 0.125 * 72 = 9).
        Real length = pixel_length_in_points / points_per_foot.
        """
        scale_ratio = self._parse_scale_ratio(scale_str) if scale_str else 0.0
        if scale_ratio <= 0:
            return  # Can't compute without scale

        points_per_foot = scale_ratio * 72.0
        if points_per_foot <= 0:
            return

        for det in detections:
            # Only compute if Gemini returned 0 or very small value
            if det.get("real_length_ft", 0) > 0.1:
                continue

            page_idx = det.get("page", 0)
            if page_idx >= len(self.doc):
                continue

            page = self.doc[page_idx]
            pw, ph = page.rect.width, page.rect.height

            # Convert normalized coords to PDF points using crop rect
            crop_raw = page_crops.get(page_idx)
            if crop_raw:
                cx0, cy0, cx1, cy1 = crop_raw
                crop_w = cx1 - cx0
                crop_h = cy1 - cy0
            else:
                cx0, cy0 = 0.0, 0.0
                crop_w, crop_h = pw, ph

            x1 = det["x1"] * crop_w + cx0
            y1 = det["y1"] * crop_h + cy0
            x2 = det["x2"] * crop_w + cx0
            y2 = det["y2"] * crop_h + cy0

            pixel_length = ((x2 - x1)**2 + (y2 - y1)**2)**0.5
            real_ft = pixel_length / points_per_foot

            # Sanity check: skip if > 200ft or < 0.1ft
            if 0.1 <= real_ft <= 200.0:
                det["real_length_ft"] = round(real_ft, 2)

    # ─── Orchestration ────────────────────────────────────────────────────────

    def analyze_plan(self, user_notes: str = "", user_context: dict = None) -> dict:
        """3-call pipeline: legend -> localize+detect per page (parallelized)."""
        if not self.doc:
            raise RuntimeError("Call ingest_pdf first")

        self._errors = []
        self._page_vectors = {}
        self._snap_trees = {}

        # Call 1: Extract legend from page 0 (or use engineer-provided codes)
        legend_data = None
        if user_context and user_context.get("wall_codes"):
            # Engineer provided wall codes — use directly, skip legend extraction
            codes = [c.strip() for c in user_context["wall_codes"].split(",")]
            legend_data = {
                "wall_types": [{"code": c.split("=")[0].strip(),
                               "description": c.split("=")[1].strip() if "=" in c else c,
                               "category": "wall", "material": "", "thickness_in": 0}
                              for c in codes],
                "scale": user_context.get("scale"),
                "floor_label": None,
                "orientation": None,
            }
        else:
            legend_data = self._extract_legend(0)
        self._legend_data = legend_data

        # Calls 2+3 per page: localize + detect (parallelized)
        page_results = [None] * len(self.doc)
        with ThreadPoolExecutor(max_workers=min(len(self.doc), 5)) as ex:
            futures = {
                ex.submit(self._analyze_single_page, i, user_notes, legend_data, user_context): i
                for i in range(len(self.doc))
            }
            for f in as_completed(futures):
                idx = futures[f]
                try:
                    page_results[idx] = f.result()
                except Exception as e:
                    self._errors.append({
                        "page": idx,
                        "error": f"Page analysis failed: {e}",
                        "stage": "page",
                    })

        all_wall_types = []
        all_detections = []
        page_crops = {}
        detections_per_page = {}
        scale = legend_data.get("scale") if legend_data else None

        # Collect known wall codes for validation
        known_codes = []
        if legend_data and legend_data.get("wall_types"):
            known_codes = [wt.get("code", "") for wt in legend_data["wall_types"]]

        for page_idx, result in enumerate(page_results):
            if result is None:
                detections_per_page[page_idx] = 0
                continue

            if result.get("crop_rect"):
                page_crops[page_idx] = result["crop_rect"]

            detections = result.get("detections", [])

            # Legend-page heuristic: skip page if >80% ROOM/DIM detections
            if page_idx == 0 and len(detections) > 5:
                room_dim_count = sum(
                    1 for d in detections
                    if d.get("category") in ("room", "dimension")
                    or d.get("type_code") in ("ROOM", "DIM")
                )
                if room_dim_count / len(detections) > 0.80:
                    detections = []

            # Phase 4: Validate detections
            detections = self._validate_detections(
                detections, page_idx,
                scale_str=result.get("scale") or (scale if scale else None),
                known_wall_codes=known_codes if known_codes else None,
            )

            detections_per_page[page_idx] = len(detections)

            for wt in result.get("wall_types", []):
                if wt not in all_wall_types:
                    all_wall_types.append(wt)

            for det in detections:
                all_detections.append(det)

            if result.get("scale") and not scale:
                scale = result["scale"]

        # Compute real_length_ft from coordinates + scale for detections where Gemini returned 0
        self._compute_real_lengths(all_detections, page_crops, scale)

        self.analysis = {
            "wall_types": all_wall_types,
            "scale": scale,
            "detections": all_detections,
            "page_crops": page_crops,
            "page_vectors": {k: len(v) for k, v in self._page_vectors.items()},
            "detections_per_page": detections_per_page,
            "errors": self._errors,
            "legend_data": legend_data,
        }
        return self.analysis

    # ─── Annotation ───────────────────────────────────────────────────────────

    def annotate_pdf(self, output_path: str) -> str:
        """Draw colored element overlays on the PDF."""
        if not self.analysis or not self.doc:
            raise RuntimeError("Call analyze_plan first")

        doc = fitz.open(self.pdf_path)
        page_crops = self.analysis.get("page_crops", {})

        for det in self.analysis["detections"]:
            page_idx = det.get("page", 0)
            if page_idx >= len(doc):
                continue
            page = doc[page_idx]
            pw, ph = page.rect.width, page.rect.height

            # Get crop rect used when this page was sent to Gemini
            crop_raw = page_crops.get(page_idx)
            if crop_raw:
                cx0, cy0, cx1, cy1 = crop_raw
                crop_w = cx1 - cx0
                crop_h = cy1 - cy0
            else:
                cx0, cy0 = 0.0, 0.0
                crop_w, crop_h = pw, ph

            raw_x1, raw_y1 = det.get("x1", 0), det.get("y1", 0)
            raw_x2, raw_y2 = det.get("x2", 0), det.get("y2", 0)

            if _is_percentage_coords(det):
                x1 = raw_x1 * crop_w + cx0
                y1 = raw_y1 * crop_h + cy0
                x2 = raw_x2 * crop_w + cx0
                y2 = raw_y2 * crop_h + cy0
            else:
                zoom = 3.0
                x1 = raw_x1 / zoom + cx0
                y1 = raw_y1 / zoom + cy0
                x2 = raw_x2 / zoom + cx0
                y2 = raw_y2 / zoom + cy0

            # Clamp to page bounds
            x1, x2 = max(0.0, min(x1, pw)), max(0.0, min(x2, pw))
            y1, y2 = max(0.0, min(y1, ph)), max(0.0, min(y2, ph))

            # Phase 1: Snap to nearest PDF vector line
            snap_data = self._snap_trees.get(page_idx)
            if snap_data:
                tree, endpoints = snap_data
                x1, y1, x2, y2 = self._snap_to_vectors(x1, y1, x2, y2, tree, endpoints)

            # Bleed line filter: skip lines that span most of the page and touch an edge
            if _is_bleed_line(x1, y1, x2, y2, pw, ph):
                continue

            # Area-ratio filter: skip room-zone bounding boxes
            bbox_area = abs(x2 - x1) * abs(y2 - y1)
            if pw * ph > 0 and bbox_area / (pw * ph) > 0.15:
                continue

            # Length sanity
            real_len = det.get("real_length_ft", 0.0)
            if real_len > 100 and det.get("category") not in ("structural",):
                continue
            if 0 < real_len < 0.5:
                continue

            type_code = det.get("type_code", "UNKNOWN")
            category = det.get("category", "wall")

            # ROOM/DIM: text label only, no line drawn
            if category in ("room", "dimension") or type_code in ("ROOM", "DIM"):
                label_color = COLOR_MAP.get(type_code) or CATEGORY_COLOR.get(category, FLAGGED_COLOR)
                page.insert_text(
                    fitz.Point((x1 + x2) / 2, (y1 + y2) / 2),
                    det.get("label", type_code),
                    fontsize=7,
                    color=label_color,
                )
                continue

            # Centerline snap: force horizontal/vertical for near-axis lines
            w = abs(x2 - x1)
            h = abs(y2 - y1)
            if w > h * 2:
                mid_y = (y1 + y2) / 2
                draw_x1, draw_y1 = min(x1, x2), mid_y
                draw_x2, draw_y2 = max(x1, x2), mid_y
            elif h > w * 2:
                mid_x = (x1 + x2) / 2
                draw_x1, draw_y1 = mid_x, min(y1, y2)
                draw_x2, draw_y2 = mid_x, max(y1, y2)
            else:
                draw_x1, draw_y1 = x1, y1
                draw_x2, draw_y2 = x2, y2

            confidence = det.get("confidence", 0.0)
            is_flagged = confidence < CONFIDENCE_THRESHOLD or type_code == "UNKNOWN"
            is_very_uncertain = confidence < 0.45 or type_code == "UNKNOWN"

            color = COLOR_MAP.get(type_code) or CATEGORY_COLOR.get(category, FLAGGED_COLOR)
            draw_color = FLAGGED_COLOR if is_flagged else color

            shape = page.new_shape()
            shape.draw_line(fitz.Point(draw_x1, draw_y1), fitz.Point(draw_x2, draw_y2))
            shape.finish(color=draw_color, width=3.0, stroke_opacity=0.85)
            shape.commit()

            label = type_code if not is_flagged else f"{type_code}?"
            mid_x = (draw_x1 + draw_x2) / 2
            mid_y = (draw_y1 + draw_y2) / 2
            page.insert_text(
                fitz.Point(mid_x + 2, mid_y - 4),
                label,
                fontsize=6,
                color=draw_color,
            )

            if is_very_uncertain:
                radius = max(max(w, h) / 2 + 8, 12)
                rect = fitz.Rect(mid_x - radius, mid_y - radius, mid_x + radius, mid_y + radius)
                shape2 = page.new_shape()
                shape2.draw_oval(rect)
                shape2.finish(color=FLAGGED_COLOR, width=1.0, stroke_opacity=0.6)
                shape2.commit()

        doc.save(output_path)
        doc.close()
        return output_path

    # ─── Reporting ────────────────────────────────────────────────────────────

    def generate_report(self) -> dict:
        """Generate material quantity report from analysis."""
        if not self.analysis:
            raise RuntimeError("Call analyze_plan first")

        detections = self.analysis["detections"]
        floor_data = {}
        flagged_count = 0
        total_footage = 0.0

        for det in detections:
            page = det.get("page", 0)
            type_code = det.get("type_code", "UNKNOWN")
            confidence = det.get("confidence", 0.0)
            length_ft = det.get("real_length_ft", 0.0)

            if confidence < CONFIDENCE_THRESHOLD or type_code == "UNKNOWN":
                flagged_count += 1

            floor_key = f"Floor {page + 1}"
            if floor_key not in floor_data:
                floor_data[floor_key] = {}
            if type_code not in floor_data[floor_key]:
                floor_data[floor_key][type_code] = {"linear_ft": 0.0, "segment_count": 0}

            floor_data[floor_key][type_code]["linear_ft"] += length_ft
            floor_data[floor_key][type_code]["segment_count"] += 1
            total_footage += length_ft

        return {
            "wall_types": self.analysis["wall_types"],
            "scale": self.analysis["scale"],
            "floors": floor_data,
            "total_segments": len(detections),
            "total_linear_ft": round(total_footage, 2),
            "flagged_count": flagged_count,
            "detections_per_page": self.analysis.get("detections_per_page", {}),
            "errors": self.analysis.get("errors", []),
            "legend_data": self.analysis.get("legend_data"),
        }

    # ─── Excel export ─────────────────────────────────────────────────────────

    def export_excel(self, report: dict, output_path: str) -> str:
        """Export takeoff report to Excel (.xlsx)."""
        import openpyxl
        from openpyxl.styles import Font, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Takeoff Report"

        headers = ["Floor", "Type Code", "Description", "Linear Footage (ft)", "Segment Count"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        type_desc = {
            wt.get("code", ""): wt.get("description", "")
            for wt in report.get("wall_types", [])
        }

        row = 2
        total_footage = 0.0
        total_segments = 0
        for floor, types in sorted(report.get("floors", {}).items()):
            for code, info in sorted(types.items()):
                ws.cell(row=row, column=1, value=floor)
                ws.cell(row=row, column=2, value=code)
                ws.cell(row=row, column=3, value=type_desc.get(code, ""))
                ws.cell(row=row, column=4, value=round(info["linear_ft"], 2))
                ws.cell(row=row, column=5, value=info["segment_count"])
                total_footage += info["linear_ft"]
                total_segments += info["segment_count"]
                row += 1

        ws.cell(row=row, column=1, value="TOTAL")
        ws.cell(row=row, column=4, value=round(total_footage, 2))
        ws.cell(row=row, column=5, value=total_segments)
        for col in range(1, 6):
            ws.cell(row=row, column=col).font = Font(bold=True)

        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=0)
            ws.column_dimensions[col[0].column_letter].width = max(max_len + 4, 12)

        wb.save(output_path)
        return output_path

    # ─── Full pipeline ────────────────────────────────────────────────────────

    def run(self, pdf_path: str, output_pdf_path: str = None,
            user_notes: str = "", user_context: dict = None) -> dict:
        """Full pipeline: ingest -> analyze -> annotate -> report."""
        info = self.ingest_pdf(pdf_path)

        if output_pdf_path is None:
            base = os.path.splitext(pdf_path)[0]
            output_pdf_path = f"{base}_annotated.pdf"

        analysis = self.analyze_plan(user_notes=user_notes, user_context=user_context)
        self.annotate_pdf(output_pdf_path)
        report = self.generate_report()

        return {
            "pdf_info": info,
            "analysis": analysis,
            "annotated_pdf": output_pdf_path,
            "report": report,
        }
